#!/usr/bin/env python3
"""
beleg_excel.py — Excel-Verwaltung für Belege.

Erstellt/aktualisiert eine Excel-Datei mit:
- Einem Tab pro Jahr (z.B. "2026", "2027")
- Spalten: Datum, Kategorie, Rechnungssteller, Bezeichnung, Netto, MwSt, Brutto
- Neue Belege werden ans Ende des jeweiligen Jahrestabs angehängt
"""

import os
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from config import EXCEL_DATEI

# Excel-Import
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("FEHLER: openpyxl nicht installiert. Bitte: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


KOPFZEILE = ["Datum", "Kategorie", "Rechnungssteller", "Bezeichnung",
             "Netto (EUR)", "MwSt-Satz", "MwSt (EUR)", "Brutto (EUR)",
             "Belegnummer", "Belegdatei"]

SPALTENBREITEN = [14, 28, 28, 45, 14, 12, 14, 14, 18, 55]

KOPF_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
KOPF_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
KOPF_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
ZELLEN_ALIGNMENT = Alignment(vertical="top", wrap_text=False)
BORDER_THIN = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
EURO_FONT = Font(name="Calibri", size=11)

JAHRES_FILLS = {
    0: None,  # Standard
    1: PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # hellgrün
    2: PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),  # hellblau
    3: PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # hellorange
}


def _excel_oeffnen(pfad=None):
    """Öffnet die Excel-Datei (oder erstellt sie neu)."""
    p = pfad or EXCEL_DATEI
    path = Path(p)
    if path.exists():
        wb = load_workbook(path)
        print(f"📊 Excel geöffnet: {p}")
    else:
        wb = Workbook()
        # Standard-Blatt entfernen
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
        print(f"📊 Neue Excel-Datei erstellt: {p}")

    return wb


def _tab_sichern(ws, jahr: str):
    """Stellt sicher, dass der Tab für das Jahr existiert und formatiert ist."""
    ws.title = jahr
    # Kopfzeile
    if ws.max_row == 1 or ws.cell(1, 1).value != "Datum":
        for col_idx, (kopf, breite) in enumerate(zip(KOPFZEILE, SPALTENBREITEN), 1):
            cell = ws.cell(1, col_idx, kopf)
            cell.font = KOPF_FONT
            cell.fill = KOPF_FILL
            cell.alignment = KOPF_ALIGNMENT
            cell.border = BORDER_THIN
            ws.column_dimensions[get_column_letter(col_idx)].width = breite
        # Kopfzeile fixieren
        ws.freeze_panes = "A2"


def _naechste_zeile(ws):
    """Findet die nächste freie Zeile im Tab."""
    for row in range(2, ws.max_row + 2):
        if ws.cell(row, 1).value is None:
            return row
    return ws.max_row + 1


def _beleg_exists(ws, datum: str, rechnungssteller: str, brutto: float) -> bool:
    """Prüft, ob ein Beleg bereits existiert (Dublettencheck)."""
    for row in range(2, ws.max_row + 1):
        d = ws.cell(row, 1).value
        r = ws.cell(row, 3).value
        b = ws.cell(row, 8).value
        if d == datum and r == rechnungssteller and b == brutto:
            return True
    return False


def beleg_hinzufuegen(beleg: dict, ueberschreiben: bool = False, excel_path: str = None) -> str:
    """
    Fügt einen Beleg zur Excel-Datei hinzu.
    - beleg: Dict mit den Feldern
    - ueberschreiben: Wenn True, überschreibe Dubletten
    - excel_path: Optional anderer Pfad (z.B. für Nicht-Finanzdokumente)
    Gibt eine Statusmeldung zurück.
    """
    pfad = excel_path or EXCEL_DATEI
    jahr = beleg.get("datum", "")[:4] or str(datetime.now().year)
    wb = _excel_oeffnen(pfad)

    # Tab für Jahreszahl
    if jahr not in wb.sheetnames:
        ws = wb.create_sheet(jahr)
        _tab_sichern(ws, jahr)
    else:
        ws = wb[jahr]

    # Dublettencheck
    datum = str(beleg.get("datum", ""))
    rechnungssteller = str(beleg.get("rechnungssteller", ""))
    brutto = float(beleg.get("brutto", 0))

    if _beleg_exists(ws, datum, rechnungssteller, brutto) and not ueberschreiben:
        return f"⏭️  Beleg existiert bereits: {datum} — {rechnungssteller} ({brutto:.2f} EUR)"

    zeile = _naechste_zeile(ws)

    # Daten schreiben
    werte = [
        datum,
        str(beleg.get("kategorie", "")),
        rechnungssteller,
        str(beleg.get("bezeichnung", "")),
        float(beleg.get("netto", 0)),
        str(beleg.get("mwst_satz", "")) + "%",
        float(beleg.get("mwst_betrag", 0)),
        brutto,
        str(beleg.get("belegnummer", "")),
        str(beleg.get("_dateiname", "")),
    ]

    for col_idx, wert in enumerate(werte, 1):
        cell = ws.cell(zeile, col_idx, wert)
        cell.font = EURO_FONT
        cell.alignment = ZELLEN_ALIGNMENT
        cell.border = BORDER_THIN

        # Euro-Spalten formatieren
        if col_idx in (5, 7, 8):  # Netto, MwSt, Brutto
            cell.number_format = '#,##0.00'

    wb.save(pfad)
    print(f"✅ Beleg in Excel geschrieben: {pfad} — Tab '{jahr}', Zeile {zeile}")
    return f"✅ In Excel gespeichert: Tab '{jahr}'"


def excel_summary() -> str:
    """Erzeugt eine Zusammenfassung der Excel-Inhalte."""
    if not os.path.exists(EXCEL_DATEI):
        return "📊 Keine Excel-Datei vorhanden."

    wb = load_workbook(EXCEL_DATEI)
    zeilen = [f"📊 **Belege-Excel — {EXCEL_DATEI}**\n"]

    for tab_name in wb.sheetnames:
        ws = wb[tab_name]
        if ws.max_row <= 1:
            continue

        # Anzahl Belege
        anzahl = ws.max_row - 1
        # Summe Brutto
        summe = 0.0
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row, 8).value
            if val and isinstance(val, (int, float)):
                summe += val

        zeilen.append(f"- **{tab_name}**: {anzahl} Belege, {summe:,.2f} EUR Brutto")

    wb.close()
    return "\n".join(zeilen)


# ─── CLI ─────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import json

    if len(sys.argv) > 1 and sys.argv[1] == "--summary":
        print(excel_summary())
    elif len(sys.argv) > 1 and sys.argv[1] == "--add":
        beleg_json = sys.stdin.read()
        beleg = json.loads(beleg_json)
        msg = beleg_hinzufuegen(beleg)
        print(msg)
    else:
        print("Verwendung:")
        print("  python beleg_excel.py --summary           # Zusammenfassung")
        print("  echo '<json>' | python beleg_excel.py --add  # Beleg hinzufügen")
        print(f"\nExcel-Datei: {EXCEL_DATEI}")